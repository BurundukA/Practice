import openpyxl


def extract_text_between_commas(filename, link_count, start_row=2):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    extracted_text = []

    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        if row[0] and isinstance(row[0], str):
            elements = row[0].split(',')
            if len(elements) >= 3:
                text_between_commas = elements[2]
                extracted_text.append(text_between_commas)
                if len(extracted_text) == link_count:
                    break

    workbook.close()
    return extracted_text


def extract_text_before_comma(filename, link_count, start_row=2):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    extracted_text = []

    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        if row[0] and isinstance(row[0], str):
            elements = row[0].split(',')
            if len(elements) >= 1:
                text_before_comma = elements[0].strip()
                extracted_text.append(text_before_comma)
                if len(extracted_text) == link_count:
                    break

    workbook.close()
    return extracted_text